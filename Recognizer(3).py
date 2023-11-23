import cv2
import openpyxl
import List
from datetime import datetime

# Load face recognition model
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read("Trainer/Trainer.yml")
cascadePath = cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
faceCascade = cv2.CascadeClassifier(cascadePath)

font = cv2.FONT_HERSHEY_SIMPLEX

# Load names from List module
names = [getattr(List, attr) for attr in dir(List) if attr.startswith("name_")]

# Initialize webcam
cam = cv2.VideoCapture(0)
cam.set(3, 640)
cam.set(4, 480)

minW = 0.1 * cam.get(3)
minH = 0.1 * cam.get(4)

while True:
    ret, img = cam.read()
    img = cv2.flip(img, 1)

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    faces = faceCascade.detectMultiScale(
        gray, scaleFactor=1.2, minNeighbors=5, minSize=(int(minW), int(minH)),
    )

    for (x, y, w, h) in faces:
        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)

        id, confidence = recognizer.predict(gray[y:y + h, x:x + w])

        if confidence < 150:
            id = names[id]
            confidence = "          {0}%".format(round(confidence))
        else:
            id = "Unknown"
            confidence = "          0%"

        cv2.putText(img, str(id), (x + 5, y - 5), font, 1, (255, 255, 255), 2)
        cv2.putText(img, str(confidence), (x + 5, y + h - 5), font, 1, (255, 255, 255), 2)


    cv2.imshow("Nhan dien khuon mat", img)

    k = cv2.waitKey(5) & 0xff
    if k == 27:
        break

# Create Excel file
current_time = datetime.now().strftime("%Y-%m-%d")
current_date = datetime.now().strftime("%H:%M:%S")
filename = f"{id}_cham_cong.xlsx"

def write_to_excel(filename, data):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Danh sách khuôn mặt"
        sheet.cell(row=1, column=1, value="Date")
        sheet.cell(row=1, column=2, value="Time")
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 15

    # Find the next available row
    next_row = sheet.max_row + 1

    # Write data to the sheet
    sheet.cell(row=next_row, column=1, value=data['date'])
    sheet.cell(row=next_row, column=2, value=data['time'])

    # Save the Excel file
    workbook.save(filename)
    workbook.close()

data_entry = {'date': current_date, 'time': current_time}
write_to_excel(filename, data_entry)

cam.release()
cv2.destroyAllWindows()
