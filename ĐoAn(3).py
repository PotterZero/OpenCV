import cv2
import List
import numpy as np
import os
import time
from datetime import datetime
import serial
import openpyxl

recognizer = cv2.face.LBPHFaceRecognizer.create()
recognizer.read("Trainer/Trainer.yml")
cascadePath = cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
faceCascade = cv2.CascadeClassifier(cascadePath)

font = cv2.FONT_HERSHEY_SIMPLEX
names = [getattr(List, attr) for attr in dir(List) if attr.startswith("name_")]

cam = cv2.VideoCapture(0)
cam.set(3, 640)
cam.set(4, 480)

minW = 0.1 * cam.get(3)
minH = 0.1 * cam.get(4)

# Khởi tạo biến id trước vòng lặp
id = "Unknown"

# Đọc nội dung từ file chỉ một lần ngoài vòng lặp
with open("Encoding.py", "r", encoding="utf-8") as file:
    content = file.read()
executed_once = False

while True:
    ret, img = cam.read()
    img = cv2.flip(img, 1)

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    faces = faceCascade.detectMultiScale(
        gray,
        scaleFactor=1.2,
        minNeighbors=5,
        minSize=(int(minW), int(minH)),
    )

    for (x, y, w, h) in faces:
        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)

        # Sử dụng biến local_id để giữ giá trị của id trong vòng lặp
        local_id, confidence = recognizer.predict(gray[y:y + h, x:x + w])

        if confidence < 150:
            local_id = names[local_id]
            confidence = "          {0}%".format(round(confidence))
            if not executed_once:
                    exec(content)
                    executed_once = True


        else:
            local_id = "Unknown"
            confidence = "          0%"

        cv2.putText(img, str(local_id), (x + 5, y - 5), font, 1, (255, 255, 255), 2)
        cv2.putText(img, str(confidence), (x + 5, y - 5), font, 1, (255, 255, 255), 2)

        # Gán giá trị của local_id cho biến id
        id = local_id

    cv2.imshow("Nhan dien khuon mat", img)

    k = cv2.waitKey(5) & 0xff
    if k == 27:
        break

# Di chuyển dòng này ra khỏi vòng lặp
cam.release()
cv2.destroyAllWindows()
print("ID :", id)

# Create Excel file
current_time = datetime.now().strftime("%Y-%m-%d")
current_date = datetime.now().strftime("%H:%M:%S")
filename = f"{id}_cham_cong.xlsx"

def write_to_excel(filename, data):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Danh sách khuôn mặt"
        sheet.cell(row=1, column=1, value="Date")
        sheet.cell(row=1, column=2, value="Time")
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 15

    # Find the next available row
    next_row = sheet.max_row + 1

    # Write data to the sheet
    sheet.cell(row=next_row, column=1, value=data['date'])
    sheet.cell(row=next_row, column=2, value=data['time'])

    # Save the Excel file
    workbook.save(filename)
    workbook.close()

data_entry = {'date': current_date, 'time': current_time}
write_to_excel(filename, data_entry)

cam.release()
cv2.destroyAllWindows()

